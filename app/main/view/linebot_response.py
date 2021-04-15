from os import access
from app.main.service import ret
import json
from app.main.constant import LineConstant
import requests as urllib_requests
import platform
from app.main.model.user import sdUser
from app.main import db
from openpyxl import load_workbook, Workbook
from pathlib import Path


def check_line_user(payload) -> str:
    temp = payload.get("events")[0]
    replytoken = temp["replyToken"]
    msg_text = temp["message"]["text"]  # HINT name
    user_id = temp["source"]["userId"]

    invitation_url = generate_url(user_id)
    search_res = sdUser.search(user_id)
    if search_res is None:
        add_line_user_to_db(search_res, msg_text, user_id)
    return invitation_url, replytoken


def generate_url(user_id: str):
    if platform.system() == "Windows":
        invitation_url = (f"https://notify-bot.line.me/oauth/authorize?response_type=code&scope=notify&"
                          f"response_mode=form_post&client_id={LineConstant.NOTIFY.get('CLIENT_ID')}"
                          f"&redirect_uri={LineConstant.NOTIFY.get('local_URI')}"
                          f"&state={user_id}")
    else:  # Linux
        invitation_url = (f"https://notify-bot.line.me/oauth/authorize?response_type=code&scope=notify&"
                          f"response_mode=form_post&client_id={LineConstant.NOTIFY.get('CLIENT_ID')}"
                          f"&redirect_uri={LineConstant.NOTIFY.get('remote_URI')}"
                          f"&state={user_id}")
    return invitation_url


def add_line_user_to_db(search_res, msg_text, user_id):
    if search_res is None:
        try:
            obj = sdUser().add(msg_text, user_id)
            db.session.add(obj)
            db.session.commit()
        except Exception as e:
            print((f"failed to update data to SQL: {str(e)}"))
            # logger.error(f"failed to update data to SQL: {str(e)}")
            db.session.rollback()
            raise
        finally:
            db.session.close()


def retrieve_notify_token_from_callback(request):
    code = request.form.get('code')
    user_id = request.form.get('state')  # 我將state故意設定為資料庫中對應的user_id，用來統整messaging API and Notify的使用者
    files = {
        "grant_type": "authorization_code",
        "client_id": LineConstant.NOTIFY.get('CLIENT_ID'),
        "client_secret": LineConstant.NOTIFY.get('SECRET'),
        "code": code
    }
    if platform.system() == "Windows":
        files.update({"redirect_uri": LineConstant.NOTIFY.get('local_URI')})
    else:  # Linux
        files.update({"redirect_uri": LineConstant.NOTIFY.get('remote_URI')})
    print("------------")
    print("code",  code)  # for debug
    print("------------")

    # HINT magic method, 從網路上抄的，還不確定是否一定要這樣寫 https://stackoverflow.com/questions/20759981/python-trying-to-post-form-using-requests by atupal#
    session = urllib_requests.Session()
    result = session.post(
        LineConstant.OFFICIAL_OAUTH_API,
        headers={'User-Agent': 'Mozilla/5.0'},
        params=files
    )
    output = json.loads(result.text)
    print("------------")
    print(f"output: {output}")
    print("------------")
    access_token = output.get("access_token")
    # append_notify_token(user_id, access_token)
    try:
        obj = sdUser().update(user_id, access_token)
        db.session.add(obj)
        db.session.commit()
    except Exception as e:
        print(f"failed to update user: {e}")


def webhook_message_checker(payload):
    try:
        if payload.get("events"):
            temp = payload.get("events")[0]
            if temp["message"]["type"] == "text":
                print("------ type is text ------")
                return "text"
            elif temp["message"]["type"] == "image":
                print("------ type is image ------")
                return "image"
            elif temp["message"]["type"] == "file":
                print("------ type is file ------")
                return "file"
            else:
                print("------ type is unknown ------")
                return False
        else:
            return False
    except Exception as e:
        print(f"-------invalid payload: {e}-------")
        return False


def excel_handler():
    input_file = "demo_doc.xlsx"
    wb = load_workbook(Path.cwd() / "downloads/" / input_file)
    ws = wb.active

    # HINT row = 1,2,3,4... in Excel; col = A,B,C... in Excel
    for row in ws.iter_rows(min_row=3, max_col=10, max_row=500, values_only=True):
        pass
